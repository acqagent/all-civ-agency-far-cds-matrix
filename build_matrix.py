"""Build a per-agency provision/clause matrix.

Master row source: WarU "Provision & Clause Matrix" (042922026).xlsx
  - Includes every 52.* entry (pre-RFO + retained + RFO additions = ~867 rows)
  - Carries P/C, prescription FAR ref, prescription text, effective date
  - "RFO" column distinguishes Removed-by-RFO vs Retained vs Added-by-RFO

Agency dates are rolled up at the FAR Part level from the existing HHS-format file
(which itself was derived from far_class_deviations-2026-04-27.xlsx in this repo).

Output:
  - far_provisions_clauses_matrix.xlsx
    One README + one tab per agency. Each tab has every 52.* entry as a row.
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

REPO = Path(r"C:\Users\tijki\rfo-deviations-repo")
DOWNLOADS = Path(r"C:\Users\tijki\Downloads")
DESKTOP = Path(r"C:\Users\tijki\Desktop")

WARU_XLSX = DESKTOP / "WarU Provision & Clause Matrix (042922026).xlsx"
HHS_FORMAT_XLSX = DOWNLOADS / "far_class_deviations_hhs_format.xlsx"
CORPUS_PROVISIONS_XLSX = REPO / "far_provisions_clauses-2026-04-27.xlsx"
OUTPUT_XLSX = DOWNLOADS / "far_provisions_clauses_matrix.xlsx"

CLAUSE_NUM_RE = re.compile(r"^52\.(\d{3})-(\d+[A-Za-z]?)$")

# Title corrections that survive Unicode normalization.
# Sourced from acquisition.gov post-RFO Part 52 (corpus baseline).
TITLE_OVERRIDES: dict[str, str] = {
    "52.209-14": "Reserve Officer Training Corps and Military Recruiting on Campus",
    "52.204-7": "System for Award Management—Registration",
    "52.212-4": "Terms and Conditions—Commercial Products and Commercial Services",
    "52.213-4": "Terms and Conditions—Simplified Acquisitions (Noncommercial)",
    "52.219-18": "Notification of Competition Limited to Eligible 8(a) Participants",
    "52.223-2": "Reporting of Biobased Products Under Service and Construction Contracts",
    "52.223-11": "Ozone-Depleting Substances",
}

# Type corrections (Provision <-> Clause) where WarU master is wrong.
TYPE_OVERRIDES: dict[str, str] = {
    "52.241-1": "Clause",
}

# Per-(agency, FAR Part) effective-date overrides for memos that bundle many
# Parts under one PDF but specify a different effective date per Part. The
# corpus only carries one date per memo, so it can't represent this on its own.
#
# DOC PM 2026-06 REVISED (signed 2026-04-22) consolidates 14 prior PMs and sets
# per-Part effective dates explicitly in the body. Most Parts now share the
# universal "January 15, 2026" post-RFO transition date; the early-adoption
# Parts retain their original 2025 effective dates.
AGENCY_PART_DATE_OVERRIDES: dict[tuple[str, int], str] = {
    ("DOC", 1):  "May 2025",
    ("DOC", 6):  "Jul 2025",
    ("DOC", 10): "May 2025",
    ("DOC", 11): "Jul 2025",
    ("DOC", 18): "Jun 2025",
    ("DOC", 29): "Aug 2025",
    ("DOC", 31): "Aug 2025",
    ("DOC", 34): "May 2025",
    ("DOC", 39): "Jun 2025",
    ("DOC", 43): "Jun 2025",
    # All other DOC Parts: "Jan 2026" (universal post-RFO transition).
    **{("DOC", p): "Jan 2026" for p in (
        2, 3, 4, 5, 7, 8, 9, 12, 13, 14, 15, 16, 17, 19, 22, 23, 24, 25, 26,
        27, 28, 30, 32, 33, 35, 36, 37, 38, 40, 41, 42, 44, 45, 46, 47, 48,
        49, 50, 51, 52, 53,
    )},
}


def part_from_number(number: str) -> int | None:
    """52.203-3 -> Part 3.  52.252-1 -> Part 52."""
    base = number.split(",")[0].strip()
    m = CLAUSE_NUM_RE.match(base)
    if not m:
        return None
    return int(m.group(1)) - 200


def clean_pc(value) -> str:
    """Normalize 'P or C' cells. Strip unicode noise; map to Provision/Clause."""
    if not value:
        return ""
    s = "".join(ch for ch in str(value) if unicodedata.category(ch)[0] != "C" and ch.isascii())
    s = s.strip().rstrip("-_*").upper()
    if s.startswith("P"):
        return "Provision"
    if s.startswith("C"):
        return "Clause"
    return ""


def rfo_status(value) -> str:
    if value is None or str(value).strip() == "":
        return "Retained"
    s = str(value).strip().upper()
    if "X" in s:
        return "Removed by RFO"
    if s == "ADD":
        return "Added by RFO"
    return s  # surface anything unexpected verbatim


def normalize_date(value) -> str:
    """Strip wrapping noise and return date as a clean string."""
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%b %Y")
    return str(value).strip()


def load_corpus_overrides() -> dict[str, dict]:
    """Load post-RFO authoritative effective dates and titles from the corpus.

    For Retained clauses, the corpus reflects what acquisition.gov currently shows
    in post-RFO Part 52 — newer than the WarU master in many cases (e.g. 52.227-1
    was revised Jun 2020 but WarU still carries Apr 1984).
    """
    if not CORPUS_PROVISIONS_XLSX.exists():
        return {}
    wb = openpyxl.load_workbook(CORPUS_PROVISIONS_XLSX, data_only=True, read_only=True)
    ws = wb["Provisions & Clauses"]
    out: dict[str, dict] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, title, typ, eff = row[0], row[1], row[2], row[3]
        if not num:
            continue
        out[str(num).strip()] = {
            "title": (title or "").strip(),
            "type": typ,
            "effective": (eff or "").strip() if isinstance(eff, str) else eff,
        }
    return out


def load_master() -> list[dict]:
    """Load 52.* rows from the WarU matrix, overriding stale dates/titles from corpus."""
    corpus = load_corpus_overrides()
    wb = openpyxl.load_workbook(WARU_XLSX, data_only=True)
    ws = wb["Matrix"]
    # Header row is 6; data starts at row 7.
    rows: list[dict] = []
    for r in range(7, ws.max_row + 1):
        num = ws.cell(row=r, column=2).value
        if not num or not isinstance(num, str):
            continue
        num_s = num.strip()
        if num_s.startswith("See notes"):
            continue
        if not num_s.startswith("52."):
            continue
        title = ws.cell(row=r, column=3).value or ""
        date = normalize_date(ws.cell(row=r, column=4).value)
        prescription_ref = ws.cell(row=r, column=5).value or ""
        prescription_text = ws.cell(row=r, column=6).value or ""
        pc = clean_pc(ws.cell(row=r, column=7).value)
        status = rfo_status(ws.cell(row=r, column=8).value)

        title_s = str(title).strip()
        # Corpus is authoritative for retained clause base data.
        # Don't override "Removed by RFO" rows — corpus only carries retained clauses.
        if status != "Removed by RFO" and num_s in corpus:
            c = corpus[num_s]
            if c["effective"]:
                date = normalize_date(c["effective"])
            # Only swap title if corpus title is non-empty and not a corpus-side
            # parsing truncation (clipped at first period).
            ct = c["title"]
            if ct and len(ct) > 3 and not ct.endswith((" U", " F")):
                title_s = ct
            if c["type"] in ("Provision", "Clause"):
                pc = c["type"]

        # Hand-curated overrides win over both corpus and WarU.
        if num_s in TITLE_OVERRIDES:
            title_s = TITLE_OVERRIDES[num_s]
        if num_s in TYPE_OVERRIDES:
            pc = TYPE_OVERRIDES[num_s]

        rows.append(
            {
                "number": num_s,
                "title": title_s,
                "type": pc,
                "rfo_status": status,
                "effective": date,
                "prescription_ref": str(prescription_ref).strip(),
                "prescription_text": str(prescription_text).strip(),
                "part": part_from_number(num_s),
            }
        )
    return rows


def load_agency_part_dates() -> dict[str, dict[int, dict]]:
    wb = openpyxl.load_workbook(HHS_FORMAT_XLSX, data_only=True)
    out: dict[str, dict[int, dict]] = {}
    for sname in wb.sheetnames:
        if sname == "README":
            continue
        ws = wb[sname]
        headers = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers) if h is not None}
        date_col = next((h for h in headers if h and isinstance(h, str) and h.endswith("Deviation Date")), None)
        if date_col is None:
            continue
        per_part: dict[int, dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                part = int(str(row[idx["Part"]]).strip())
            except (TypeError, ValueError):
                continue
            raw_date = str(row[idx[date_col]] or "")
            is_immediate = bool(re.search(r"\(?[Ii]mmediate\)?", raw_date))
            clean_date = re.sub(r"\s*\(?[Ii]mmediate\)?\s*", "", raw_date).strip()
            notes = row[idx["Notes"]] or ""
            entry = {
                "date": clean_date,
                "effective": "Immediate" if is_immediate else "",
                "disposition": row[idx["Disposition"]] or "",
                "notes": notes,
            }
            # Expand multi-Part deviation memos. The HHS-format file stores one
            # row per "primary" Part, but a single PDF often covers many Parts
            # (e.g. DOC's "FAR Parts 1 through 53" or USAID's explicit list).
            for p in expand_parts(part, notes):
                per_part[p] = entry
        out[sname] = per_part
    return out


def expand_parts(primary: int, notes: str) -> list[int]:
    """Return every FAR Part this deviation row applies to.

    Reads multi-Part hints from the Notes/Source text. Recognized forms:
      - Filename list: "Parts-3-17-27-45and52" -> [3,17,27,45,52]
      - Filename two-Part: "Parts-1and34" -> [1,34]
      - Filename range: "Parts-1-53" (only when exactly two numbers) -> 1..53
      - Prose range: "FAR Parts 1 through 53" / "Parts 1 to 53" -> 1..53
      - Prose list: "parts 1, 6, 10 ... and 43" -> [1,6,10,...,43]
    Always includes the primary Part. De-duped, sorted.
    """
    parts = {primary}
    text = str(notes or "")

    # Filename two-Part with "and": "Parts-1and34.pdf" -> [1, 34]
    for m in re.finditer(r"[Pp]arts?-(\d{1,2})and(\d{1,2})\.pdf", text):
        for v in (int(m[1]), int(m[2])):
            if 1 <= v <= 53:
                parts.add(v)

    # Filename list-or-range: "Parts-N-N-...-N(andN)?.pdf"
    for m in re.finditer(r"[Pp]arts?-(\d{1,2}(?:-\d{1,2})+(?:and\d{1,2})?)\.pdf", text):
        nums = [int(n) for n in re.findall(r"\d{1,2}", m.group(1))]
        if len(nums) >= 3 or "and" in m.group(1):
            # 3+ hyphen-separated numbers, or any "andN" suffix -> explicit list
            for v in nums:
                if 1 <= v <= 53:
                    parts.add(v)
        elif len(nums) == 2:
            lo, hi = sorted(nums)
            # Two-number filename: only treat as a range when the gap is wide
            # enough that an explicit list would be implausible (>= 5).
            if hi - lo >= 5 and 1 <= lo <= 53 and 1 <= hi <= 53:
                parts.update(range(lo, hi + 1))
            else:
                parts.update(nums)

    # Prose range: "Parts X through Y" or "Parts X to Y"
    for m in re.finditer(r"[Pp]arts?\s+(\d{1,2})\s+(?:through|to)\s+(\d{1,2})\b", text):
        try:
            lo, hi = int(m[1]), int(m[2])
        except ValueError:
            continue
        if hi > lo and 1 <= lo <= 53 and 1 <= hi <= 53:
            parts.update(range(lo, hi + 1))

    # Prose list: "parts 1, 6, 10, 11, 18, 29, 31, 34, 39, 43, and corresponding ..."
    prose = re.search(
        r"[Pp]arts?\s+((?:\d{1,2}(?:\s*,\s*|\s+and\s+)){2,}\d{1,2})\b", text
    )
    if prose:
        for n in re.findall(r"\d{1,2}", prose.group(1)):
            v = int(n)
            if 1 <= v <= 53:
                parts.add(v)

    return sorted(parts)


HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(color="FFFFFF", bold=True)
PROV_FILL = PatternFill("solid", fgColor="FEF3C7")
CLAUSE_FILL = PatternFill("solid", fgColor="DBEAFE")
REMOVED_FILL = PatternFill("solid", fgColor="FEE2E2")
ADDED_FILL = PatternFill("solid", fgColor="DCFCE7")
NO_DEV_FILL = PatternFill("solid", fgColor="F3F4F6")


def write_agency_sheet(wb, agency: str, master: list[dict], part_dates: dict[int, dict]) -> None:
    ws = wb.create_sheet(agency[:31])
    headers = [
        "FAR Part",
        "Number",
        "Title",
        "Type",
        "RFO Status",
        "Effective Date (52.103)",
        "Prescription FAR Ref",
        f"{agency} Deviation Date",
        "Deviation Effective",
        "Disposition",
        "Source / Notes",
        "Prescription Text",
    ]
    ws.append(headers)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "C2"

    for entry in master:
        part = entry["part"]
        dev = part_dates.get(part) if part is not None else None
        # Per-(agency, Part) override wins over the memo-level date.
        # Used for memos that bundle many Parts with different effective dates.
        override = AGENCY_PART_DATE_OVERRIDES.get((agency, part)) if part is not None else None
        # FAR 52.103(a): when a provision/clause is used with an authorized
        # deviation, the CO inserts "(DEVIATION)" after the date in the citation.
        eff = entry["effective"]
        eff_display = f"{eff} (DEVIATION)" if dev and eff else eff
        dev_date = override or (dev["date"] if dev else "")
        row = [
            part if part is not None else "",
            entry["number"],
            entry["title"],
            entry["type"],
            entry["rfo_status"],
            eff_display,
            entry["prescription_ref"],
            dev_date,
            dev["effective"] if dev else "",
            dev["disposition"] if dev else "",
            dev["notes"] if dev else "",
            entry["prescription_text"],
        ]
        ws.append(row)
        r = ws.max_row

        # Type colors
        type_cell = ws.cell(row=r, column=4)
        if entry["type"] == "Provision":
            type_cell.fill = PROV_FILL
        elif entry["type"] == "Clause":
            type_cell.fill = CLAUSE_FILL

        # RFO Status colors
        status_cell = ws.cell(row=r, column=5)
        if entry["rfo_status"] == "Removed by RFO":
            status_cell.fill = REMOVED_FILL
        elif entry["rfo_status"] == "Added by RFO":
            status_cell.fill = ADDED_FILL

        # Gray out rows where this agency has no deviation for that Part
        if not dev:
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=col)
                if not cell.fill.fgColor or cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = NO_DEV_FILL

    widths = [8, 16, 50, 11, 16, 14, 22, 22, 12, 14, 70, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for col in (3, 11, 12):
        for row_cells in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row_cells:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def write_readme(wb, master_count: int, agencies: list[str], status_counts: dict[str, int]) -> None:
    ws = wb.create_sheet("README", 0)
    lines = [
        "FAR Class Deviations - Provision/Clause Matrix",
        "",
        f"Master row count: {master_count} (every 52.* entry from WarU matrix)",
        f"Agency tabs: {len(agencies)}",
        "",
        "RFO status breakdown:",
    ] + [f"  - {k}: {v}" for k, v in status_counts.items()] + [
        "",
        "Schema (per agency tab):",
        "  - FAR Part: derived from clause number (52.203-3 -> Part 3, 52.252-1 -> Part 52)",
        "  - Number: clause/provision number, including Alternate variants where present",
        "  - Title: as published; Alternate variants carry the alternate label",
        "  - Type: Provision or Clause (yellow = provision, blue = clause)",
        "  - RFO Status: Retained / Removed by RFO / Added by RFO (red = removed, green = added)",
        "  - Effective Date (52.103): clause's effective date as published in FAR Part 52.",
        "      Per FAR 52.103(a), '(DEVIATION)' is appended when the agency has issued a",
        "      class deviation that covers this clause's parent Part.",
        "  - Prescription FAR Ref: prescribing FAR section (e.g., 3.202)",
        "  - <Agency> Deviation Date: agency's deviation date for the parent FAR Part",
        "  - Deviation Effective: 'Immediate' if the memo is effective immediately on issuance; blank otherwise",
        "  - Disposition: 'Updated' if the agency has issued a class deviation for that Part",
        "  - Source / Notes: source PDF and scope, from the agency's deviation memo",
        "  - Prescription Text: full prescription guidance from FAR (long-form)",
        "",
        "Visual cues:",
        "  - Yellow Type cell = Provision",
        "  - Blue Type cell = Clause",
        "  - Red RFO Status cell = Removed by RFO",
        "  - Green RFO Status cell = Added by RFO",
        "  - Gray row = no class deviation issued for that FAR Part by this agency",
        "",
        "Sources:",
        "  - Master:    WarU Provision & Clause Matrix (042922026).xlsx (Desktop)",
        "  - Per-agency Part-level deviation dates: far_class_deviations_hhs_format.xlsx (Downloads)",
        "  - Underlying corpus: github.com/acqagent/rfo-deviations",
        "",
        "Notes:",
        "  - Agency deviation dates are rolled up at the FAR Part level. A removed clause",
        "    in a deviated Part is shown with that Part's deviation date because adopting",
        "    the RFO Part deviation effectively retires the removed clause.",
        "  - Per-clause deviation dates would require parsing each agency's deviation PDF",
        "    and are not yet captured here.",
        "",
        "Agencies:",
        "  " + ", ".join(agencies),
    ]
    for line in lines:
        ws.append([line])
    ws.column_dimensions["A"].width = 120


def main() -> None:
    print("Loading WarU master...")
    master = load_master()
    status_counts: dict[str, int] = {}
    for e in master:
        status_counts[e["rfo_status"]] = status_counts.get(e["rfo_status"], 0) + 1
    print(f"  {len(master)} entries -> {status_counts}")

    print("Loading agency Part-level deviation dates...")
    agency_part_dates = load_agency_part_dates()
    agencies = list(agency_part_dates.keys())
    print(f"  {len(agencies)} agencies")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    write_readme(wb, len(master), agencies, status_counts)
    for agency in agencies:
        write_agency_sheet(wb, agency, master, agency_part_dates[agency])
        print(f"  wrote sheet: {agency}")

    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"\nWrote: {OUTPUT_XLSX}")


if __name__ == "__main__":
    main()
