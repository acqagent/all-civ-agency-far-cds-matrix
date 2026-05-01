"""Resolve "(Immediate)" deviation dates by reading each agency's class
deviation memo and extracting the issuance / effective date.

Source of "Immediate" rows: far_class_deviations_hhs_format.xlsx (the HHS-format
per-agency file). Source of memos: source_url field in each row's Notes column,
which points to acquisition.gov's page_file_uploads/<filename>.pdf.

Date extraction is layered:
  1. Adobe digital signature: "Date: 2025.06.11 ..."
  2. Adobe Sign block:        "(Nov 21, 2025 14:13:00 EST)"
  3. Top-of-doc memo date:    "May 27, 2025"
  4. Generic Month DD, YYYY anywhere in first ~3000 chars

Pick the earliest plausible date >= April 2025 (EO 14275 signing).
Output: a (agency, part) -> "Mon YYYY (Immediate)" map written back to
the HHS-format workbook in place.
"""

from __future__ import annotations

import re
import subprocess
import sys
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date
from pathlib import Path

import openpyxl

PDF_DIR = Path(r"C:\Users\tijki\Downloads\_rfo_pdfs")
HHS_FORMAT = Path(r"C:\Users\tijki\Downloads\far_class_deviations_hhs_format.xlsx")

MONTHS = {m: i+1 for i, m in enumerate(
    ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"])}
MONTH_FULL = {m: i+1 for i, m in enumerate(
    ["january","february","march","april","may","june","july","august",
     "september","october","november","december"])}

EARLIEST = date(2025, 4, 15)  # EO 14275 signing
EO_DATE = date(2025, 4, 15)  # referenced in every memo body — not the issuance date


def fetch(url: str, dest: Path) -> bool:
    if dest.exists() and dest.stat().st_size > 1000:
        return True
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
        with urllib.request.urlopen(req, timeout=60) as r, dest.open("wb") as f:
            f.write(r.read())
        return dest.stat().st_size > 1000
    except Exception as e:  # noqa: BLE001
        print(f"  ! fetch failed {url}: {e}", file=sys.stderr)
        return False


def pdf_text(pdf: Path) -> str:
    try:
        out = subprocess.run(
            ["pdftotext", "-layout", str(pdf), "-"],
            capture_output=True, timeout=60,
        )
        return out.stdout.decode("utf-8", errors="ignore")
    except Exception:  # noqa: BLE001
        return ""


def _parse_word_date(month_word: str, day: str, year: str) -> date | None:
    key = month_word.lower()
    mon = MONTH_FULL.get(key) or MONTHS.get(key[:3])
    if not mon:
        return None
    try:
        return date(int(year), mon, int(day))
    except ValueError:
        return None


def _ok(d: date | None) -> bool:
    return bool(d and EARLIEST <= d <= date(2027, 12, 31) and d != EO_DATE)


def find_date(text: str) -> date | None:
    """Return the best-guess issuance date for the deviation memo.

    Layered preference: digital signature > Adobe Sign block > memo-header
    date > anywhere in body (excluding EO 14275 signing date).
    """
    # 1. Adobe digital signature: "Date: 2025.06.11"
    sig_dates: list[date] = []
    for m in re.finditer(r"Date:\s*(\d{4})[.\-/](\d{1,2})[.\-/](\d{1,2})", text):
        try:
            d = date(int(m[1]), int(m[2]), int(m[3]))
        except ValueError:
            continue
        if _ok(d):
            sig_dates.append(d)
    if sig_dates:
        return min(sig_dates)

    # 2. Adobe Sign block: "(Nov 21, 2025 14:13:00 EST)"
    sign_dates: list[date] = []
    for m in re.finditer(
        r"\(([A-Za-z]{3,9})\s+(\d{1,2}),\s+(\d{4})\s+\d{1,2}:\d{2}", text
    ):
        d = _parse_word_date(m[1], m[2], m[3])
        if _ok(d):
            sign_dates.append(d)
    if sign_dates:
        return min(sign_dates)

    # 3. Memo header: scan only the chunk above the first "Background"/"Purpose"
    head = text
    boundary = re.search(r"\b(Background|Purpose|PURPOSE|BACKGROUND)\b", text)
    if boundary:
        head = text[: boundary.start()]
    head_dates: list[date] = []
    for m in re.finditer(
        r"\b([A-Z][a-z]{2,8})\s+(\d{1,2}),\s+(\d{4})\b", head
    ):
        d = _parse_word_date(m[1], m[2], m[3])
        if _ok(d):
            head_dates.append(d)
    for m in re.finditer(r"\b(\d{1,2})/(\d{1,2})/(\d{4})\b", head):
        try:
            d = date(int(m[3]), int(m[1]), int(m[2]))
        except ValueError:
            continue
        if _ok(d):
            head_dates.append(d)
    if head_dates:
        return min(head_dates)

    # 4. Body scan: any Month DD, YYYY in body that isn't the EO date.
    body_dates: list[date] = []
    for m in re.finditer(
        r"\b([A-Z][a-z]{2,8})\s+(\d{1,2}),\s+(\d{4})\b", text
    ):
        d = _parse_word_date(m[1], m[2], m[3])
        if _ok(d):
            body_dates.append(d)
    if body_dates:
        return min(body_dates)
    return None


def find_date_with_meta(pdf_path: Path, text: str) -> date | None:
    """find_date plus a fallback to PDF CreationDate metadata when the text
    layer doesn't expose a usable date (some agencies leave the signature
    in an image and never type the date in the body)."""
    d = find_date(text)
    if d:
        return d
    raw = pdf_path.read_bytes()
    candidates: list[date] = []
    # XMP first (more reliable when present)
    for m in re.finditer(rb"<xmp:CreateDate>(\d{4})-(\d{2})-(\d{2})", raw):
        try:
            candidates.append(date(int(m[1]), int(m[2]), int(m[3])))
        except ValueError:
            pass
    # Then PDF info dict /CreationDate (D:YYYYMMDDhhmmss)
    for m in re.finditer(rb"/CreationDate\s*\(D:(\d{4})(\d{2})(\d{2})", raw):
        try:
            candidates.append(date(int(m[1]), int(m[2]), int(m[3])))
        except ValueError:
            pass
    candidates = [c for c in candidates if _ok(c)]
    return min(candidates) if candidates else None


def fmt(d: date) -> str:
    return f"{d.strftime('%b %Y')} (Immediate)"


def main() -> None:
    PDF_DIR.mkdir(parents=True, exist_ok=True)

    # Collect (agency, part, url) for "(Immediate)"-only rows.
    wb = openpyxl.load_workbook(HHS_FORMAT)
    targets: list[tuple[str, str, str, int]] = []  # (agency, part, url, row_num)
    for sname in wb.sheetnames:
        if sname == "README":
            continue
        ws = wb[sname]
        headers = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers) if h is not None}
        devcol_name = next((h for h in headers
                            if isinstance(h, str) and h.endswith("Deviation Date")), None)
        if devcol_name is None:
            continue
        devcol = idx[devcol_name] + 1
        notes_col = idx.get("Notes", -1)
        for r in range(2, ws.max_row + 1):
            v = ws.cell(row=r, column=devcol).value
            if v is None:
                continue
            s = str(v).strip()
            # Match bare "(Immediate)" as well as previously-resolved
            # "Mon YYYY (Immediate)" rows so re-runs are idempotent.
            if "(Immediate)" not in s and "(immediate)" not in s.lower():
                continue
            notes = ws.cell(row=r, column=notes_col + 1).value if notes_col >= 0 else ""
            url_m = re.search(r"https?://\S+\.pdf", str(notes or ""))
            if not url_m:
                continue
            part = ws.cell(row=r, column=idx["Part"] + 1).value
            targets.append((sname, str(part), url_m.group(0), r))

    # Dedupe URLs for download.
    urls = sorted({t[2] for t in targets})
    print(f"Downloading {len(urls)} unique PDFs ({len(targets)} target rows)...")

    def fname(u: str) -> Path:
        return PDF_DIR / u.rsplit("/", 1)[-1]

    with ThreadPoolExecutor(max_workers=12) as ex:
        list(as_completed([ex.submit(fetch, u, fname(u)) for u in urls]))

    # Extract one date per URL.
    print("Extracting dates...")
    url_date: dict[str, date | None] = {}
    for u in urls:
        p = fname(u)
        if not p.exists() or p.stat().st_size < 1000:
            url_date[u] = None
            continue
        text = pdf_text(p)
        url_date[u] = find_date_with_meta(p, text)

    # Apply back to workbook.
    resolved = unresolved = 0
    for agency, part, url, r in targets:
        ws = wb[agency]
        headers = [c.value for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers) if h is not None}
        devcol_name = next((h for h in headers
                            if isinstance(h, str) and h.endswith("Deviation Date")), None)
        devcol = idx[devcol_name] + 1
        d = url_date.get(url)
        if d:
            ws.cell(row=r, column=devcol).value = fmt(d)
            resolved += 1
        else:
            unresolved += 1

    wb.save(HHS_FORMAT)
    print(f"\nResolved: {resolved}  Unresolved: {unresolved}")
    print(f"Saved: {HHS_FORMAT}")

    # Print unresolved for human follow-up.
    if unresolved:
        print("\nUnresolved (no date extracted):")
        seen = set()
        for agency, part, url, r in targets:
            if url_date.get(url) is None and url not in seen:
                print(f"  {agency:10s} Part {part:>3}  {url}")
                seen.add(url)


if __name__ == "__main__":
    main()
